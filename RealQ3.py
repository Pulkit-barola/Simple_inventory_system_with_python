import openpyxl
from openpyxl import Workbook

inventory = {
  # initial inventory is empty 

}

def add_item(item_name, quantity):
  """Adds an item to the inventory or updates the quantity if it already exists."""
  if item_name in inventory:
    inventory[item_name] += quantity
  else:
    inventory[item_name] = quantity
  print(f"{quantity} {item_name} added to inventory.")

def remove_item(item_name, quantity):
  """Removes items from the inventory. If not enough items are available, 
  removes as many as possible and prints a message."""
  if item_name in inventory:
    if inventory[item_name] >= quantity:
      inventory[item_name] -= quantity
      print(f"{quantity} {item_name} removed from inventory.")
    else:
      print(f"Not enough {item_name} in inventory. Removed {inventory[item_name]}.")
      inventory[item_name] = 0
  else:
    print(f"{item_name} not found in inventory.")

def generate_report():
  """Prints a report of the current inventory."""
  print("Inventory Report:")
  for item_name, quantity in inventory.items():
    print(f"{item_name}: {quantity}")

def export_to_excel():
  """Exports the inventory data to an Excel file."""
  wb = Workbook()
  ws = wb.active
  ws['A1'] = 'Item Name'
  ws['B1'] = 'Quantity'
  row = 2
  for item_name, quantity in inventory.items():
    ws[f'A{row}'] = item_name
    ws[f'B{row}'] = quantity
    row += 1
  wb.save('inventory.xlsx')
  print("Inventory data exported to inventory.xlsx")

while True:
  print("\nInventory System Menu:")
  print("1. Add Item")
  print("2. Remove Item")
  print("3. Generate Report")
  print("4. Export to Excel")
  print("5. Exit")

  choice = input("Enter your choice (1-5): ")

  if choice == "1":
    item_name = input("Enter item name: ")
    quantity = int(input("Enter quantity: "))
    add_item(item_name, quantity)
  elif choice == "2":
    item_name = input("Enter item name: ")
    quantity = int(input("Enter quantity: "))
    remove_item(item_name, quantity)
  elif choice == "3":
    generate_report()
  elif choice == "4":
    export_to_excel()
  elif choice == "5":
    break
  else:
    print("Invalid choice. Please try again.")

